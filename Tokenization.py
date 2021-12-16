import openpyxl

tokenize_words = []

def getting_data_from_file():
    wb = openpyxl.load_workbook("DataSet.xlsx")
    ws = wb['Form Responses 1']
    colomn = ws["B"]
    for i in range(2,len(colomn)+1):
        C = 'B' + str(i)
        D = 'C' + str(i)
        x = ws[C].value


        temp = x.replace(":"," ")
        temp2 = temp.replace("'"," ")
        temp3 = temp2.replace("।"," ")
        temp4 = temp3.replace("?"," ")
        temp5 = temp4.replace(";"," ")
        temp6 = temp5.replace("/"," ")
        temp7 = temp6.replace("!"," ")
        temp8 = temp7.replace("."," ")
        temp9 = temp8.replace("’"," ")
        temp10 = temp9.replace("ঃ"," ")
        temp11 = temp10.replace("-","")

        x = temp11

        word_list =  list(x.split(" "))
        tokenize_words = []

        for word in word_list:
            if word != '':
                tokenize_words.append(word)

        words = str(tokenize_words)
        words1 = words.replace("[","")
        words2 = words1.replace("]","")
        words3 = words2.replace("'","")
        words = words3
        ws[D].value = words

    wb.save("DataSet.xlsx")

















getting_data_from_file()































